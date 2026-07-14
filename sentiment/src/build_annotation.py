# -*- coding: utf-8 -*-
"""
Phase 1 — Build annotation sheets for two independent annotators.

From interim/cleaned.csv it creates:
    - an OVERLAP set (~450 comments) that BOTH annotators label  -> used for Cohen's Kappa
    - the remaining comments split 50/50, one half to each annotator

Outputs (in sentiment/data/annotated/):
    - annotator_A.xlsx   (overlap + half of the rest)   -> Lamia
    - annotator_B.xlsx   (overlap + other half)         -> teammate
    - overlap_ids.csv    (the shared comment_ids + AI silver labels — hidden key,
                          NOT given to annotators; used later for human-vs-AI agreement)

Each .xlsx has:
    - an "Instructions" sheet (quick legend)
    - an "Annotate" sheet with dropdown validation on polarity & emotion
    - polarity/emotion left BLANK so annotators label independently (blind);
      the AI silver labels are never shown to them.

Stratified by the AI silver EMOTION label so every emotion class (incl. the rare
ones — Grief, Prayer, Victim-blaming) is represented in the Kappa overlap.
Reproducible via SEED.
"""
from __future__ import annotations
import sys
import io
from pathlib import Path
import pandas as pd
from sklearn.model_selection import train_test_split
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
# llm_labels.csv = cleaned.csv + AI silver polarity/emotion (from assemble_llm_labels.py)
LABELS = ROOT / "data" / "annotated" / "llm_labels.csv"
OUT = ROOT / "data" / "annotated"

SEED = 42
OVERLAP_N = 450

POLARITY = ["Negative", "Neutral", "Positive"]
EMOTION = ["Anger/Condemnation", "Grief/Sympathy", "Prayer/Religious",
           "Victim-blaming", "Neutral-Other"]

# Columns the annotators see (context kept minimal; abuse_type shown but they're told to ignore it)
VIEW_COLS = ["comment_id", "platform", "abuse_type", "Comment"]
HEADERS = VIEW_COLS + ["polarity", "emotion", "notes"]


def build_xlsx(df: pd.DataFrame, path: Path, who: str) -> None:
    wb = Workbook()

    # ── Instructions sheet ──
    ins = wb.active
    ins.title = "Instructions"
    ins_lines = [
        (f"Annotation sheet — {who}", True),
        ("", False),
        ("For EACH row: pick a polarity AND an emotion from the dropdowns.", False),
        ("Label the COMMENT TEXT only. Ignore the abuse_type column (it is the video topic).", False),
        ("Label independently — do not discuss with the other annotator.", False),
        ("Do not edit/reorder comment_id. Add a note for hard cases.", False),
        ("", False),
        ("polarity:  Negative | Neutral | Positive", False),
        ("emotion:   Anger/Condemnation | Grief/Sympathy | Prayer/Religious | Victim-blaming | Neutral-Other", False),
        ("", False),
        ("Full rules: sentiment/annotation_guideline.md", False),
    ]
    for i, (txt, bold) in enumerate(ins_lines, start=1):
        c = ins.cell(row=i, column=1, value=txt)
        if bold:
            c.font = Font(bold=True, size=13)
    ins.column_dimensions["A"].width = 100

    # ── Annotate sheet ──
    ws = wb.create_sheet("Annotate")
    header_fill = PatternFill("solid", fgColor="DCE6F1")
    for j, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(bold=True)
        c.fill = header_fill
    for _, row in df.iterrows():
        ws.append([row.get(c, "") for c in VIEW_COLS] + ["", "", ""])

    # column widths + wrap on Comment
    widths = {"comment_id": 10, "platform": 10, "abuse_type": 14, "Comment": 80,
              "polarity": 16, "emotion": 20, "notes": 24}
    for j, h in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = widths.get(h, 14)
    comment_col = get_column_letter(HEADERS.index("Comment") + 1)
    for r in range(2, ws.max_row + 1):
        ws[f"{comment_col}{r}"].alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"

    # dropdown validation
    n = ws.max_row
    pol_col = get_column_letter(HEADERS.index("polarity") + 1)
    emo_col = get_column_letter(HEADERS.index("emotion") + 1)
    dv_pol = DataValidation(type="list", formula1='"%s"' % ",".join(POLARITY),
                            allow_blank=True, showDropDown=False)
    dv_emo = DataValidation(type="list", formula1='"%s"' % ",".join(EMOTION),
                            allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv_pol)
    ws.add_data_validation(dv_emo)
    dv_pol.add(f"{pol_col}2:{pol_col}{n}")
    dv_emo.add(f"{emo_col}2:{emo_col}{n}")

    wb.save(path)


def main() -> None:
    df = pd.read_csv(LABELS, dtype=str, keep_default_na=False)
    OUT.mkdir(parents=True, exist_ok=True)

    # overlap (both annotators) — stratified by AI silver emotion so every
    # emotion class (incl. rare ones) is represented in the Kappa set
    overlap, rest = train_test_split(
        df, train_size=OVERLAP_N, stratify=df["emotion"], random_state=SEED
    )
    # split the rest 50/50 — stratified by emotion too
    a_only, b_only = train_test_split(
        rest, train_size=0.5, stratify=rest["emotion"], random_state=SEED
    )

    a = pd.concat([overlap, a_only]).sample(frac=1, random_state=SEED).reset_index(drop=True)
    b = pd.concat([overlap, b_only]).sample(frac=1, random_state=SEED).reset_index(drop=True)

    build_xlsx(a, OUT / "annotator_A.xlsx", "Annotator A (Lamia)")
    build_xlsx(b, OUT / "annotator_B.xlsx", "Annotator B (teammate)")
    # hidden key (NOT for annotators): the 450 overlap ids + their AI silver labels,
    # so we can later compute human-gold vs AI-silver agreement
    overlap_key = overlap[["comment_id", "polarity", "emotion"]].rename(
        columns={"polarity": "ai_polarity", "emotion": "ai_emotion"}
    )
    overlap_key.to_csv(OUT / "overlap_ids.csv", index=False)

    print("Phase 1 sheets built:")
    print(f"  overlap (both):      {len(overlap)}")
    print(f"  annotator_A.xlsx:    {len(a)}  (overlap {len(overlap)} + A-only {len(a_only)})")
    print(f"  annotator_B.xlsx:    {len(b)}  (overlap {len(overlap)} + B-only {len(b_only)})")
    print(f"  total unique:        {df['comment_id'].nunique()}")
    print("\nOverlap emotion distribution (AI silver — what the Kappa set covers):")
    print(overlap["emotion"].value_counts().to_string())


if __name__ == "__main__":
    main()
