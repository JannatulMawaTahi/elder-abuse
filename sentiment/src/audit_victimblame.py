# -*- coding: utf-8 -*-
"""
Build a human-review sheet for the Victim-blaming class (the noisiest / highest-value class).

Why: LLM labels have errors on subtle cases — comments that condemn the ABUSER or make a
general cynical remark sometimes get mislabelled as Victim-blaming (blaming the elder/family).
This sheet lets Lamia + teammate skim every Victim-blaming comment and fix the wrong ones.

Output: sentiment/data/annotated/review/victimblame_audit.xlsx   (dropdowns, pre-filled)
        sentiment/data/annotated/review/victimblame_audit.csv    (plain backup)

How to review:
  - Read the Comment. If the emotion is RIGHT, leave `final_emotion` as is.
  - If WRONG, change `final_emotion` (and `final_polarity` if needed) from the dropdown.
  - Add a short `notes` for hard cases.
Later, reconcile with src/apply_vb_audit.py (diffs final_* vs current to patch llm_labels.csv).
"""
import sys, io
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "data" / "annotated" / "llm_labels.csv"
OUTDIR = ROOT / "data" / "annotated" / "review"
OUTDIR.mkdir(parents=True, exist_ok=True)

EMOTIONS = ["Anger/Condemnation", "Grief/Sympathy", "Prayer/Religious",
            "Victim-blaming", "Neutral-Other"]
POLARITY = ["Negative", "Neutral", "Positive"]

# Columns shown to the reviewer (order matters).
COLS = ["comment_id", "Comment", "abuse_type", "platform", "lang",
        "cur_emotion", "final_emotion", "final_polarity", "is_correct?", "notes"]


def main() -> None:
    df = pd.read_csv(SRC, dtype=str, keep_default_na=False)
    vb = df[df["emotion"] == "Victim-blaming"].copy()
    if vb.empty:
        sys.exit("No Victim-blaming rows found.")

    # longest comments first — victim-blaming tends to be argued/justified, so the
    # long ones are the most informative to eyeball early.
    vb["_len"] = vb["Comment"].str.len()
    vb = vb.sort_values("_len", ascending=False)

    review = pd.DataFrame({
        "comment_id":     vb["comment_id"].values,
        "Comment":        vb["Comment"].values,
        "abuse_type":     vb["abuse_type"].values,
        "platform":       vb["platform"].values,
        "lang":           vb["lang"].values,
        "cur_emotion":    vb["emotion"].values,           # reference (read-only)
        "final_emotion":  vb["emotion"].values,           # pre-filled; reviewer edits if wrong
        "final_polarity": vb["polarity"].values,          # pre-filled; reviewer edits if wrong
        "is_correct?":    "",                             # optional Y/N shortcut
        "notes":          "",
    })
    review.to_csv(OUTDIR / "victimblame_audit.csv", index=False, encoding="utf-8")

    # ---- styled xlsx with dropdowns ----
    wb = Workbook(); ws = wb.active; ws.title = "victimblame_audit"
    navy = PatternFill("solid", fgColor="1D3557")
    thin = Border(*(4 * (Side(style="thin", color="D9DDE2"),)))

    ws.append(COLS)
    for j, _ in enumerate(COLS, start=1):
        c = ws.cell(row=1, column=j)
        c.fill = navy; c.font = Font(bold=True, color="FFFFFF", size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")

    for _, r in review.iterrows():
        ws.append([r[c] for c in COLS])

    # dropdowns
    n = len(review) + 1
    ecol = COLS.index("final_emotion") + 1
    pcol = COLS.index("final_polarity") + 1
    icol = COLS.index("is_correct?") + 1
    dv_e = DataValidation(type="list", formula1='"%s"' % ",".join(EMOTIONS), allow_blank=False)
    dv_p = DataValidation(type="list", formula1='"%s"' % ",".join(POLARITY), allow_blank=False)
    dv_i = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    ws.add_data_validation(dv_e); ws.add_data_validation(dv_p); ws.add_data_validation(dv_i)
    for row in range(2, n + 1):
        dv_e.add(ws.cell(row=row, column=ecol))
        dv_p.add(ws.cell(row=row, column=pcol))
        dv_i.add(ws.cell(row=row, column=icol))

    widths = {"comment_id": 11, "Comment": 70, "abuse_type": 13, "platform": 10, "lang": 8,
              "cur_emotion": 17, "final_emotion": 18, "final_polarity": 15,
              "is_correct?": 11, "notes": 30}
    from openpyxl.utils import get_column_letter
    for j, col in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = widths[col]
    for row in range(2, n + 1):
        ws.cell(row=row, column=COLS.index("Comment") + 1).alignment = \
            Alignment(wrap_text=True, vertical="top")
        for j in range(1, len(COLS) + 1):
            ws.cell(row=row, column=j).border = thin
    ws.freeze_panes = "A2"

    out = OUTDIR / "victimblame_audit.xlsx"
    wb.save(out)
    print(f"Wrote {out}")
    print(f"  {len(review)} Victim-blaming comments to review "
          f"(longest-first). CSV backup alongside.")
    print("  Reviewer: fix `final_emotion` / `final_polarity` on the wrong ones, add `notes`.")


if __name__ == "__main__":
    main()
