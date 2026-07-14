# -*- coding: utf-8 -*-
"""
Phase 1 (Option 3) — Build annotation sheets for FOUR annotators (full human labelling).

Design:
  - a shared OVERLAP set (450 comments) that ALL FOUR label -> Fleiss' Kappa
  - the remaining ~1,851 comments split into 4 emotion-stratified chunks, one per annotator
  => every comment gets ≥1 human label (full human-annotated dataset); the 450 overlap gets 4.

Each annotator_{A,B,C,D}.xlsx has an "Instructions" sheet + a blind "Annotate" sheet
(dropdowns on polarity & emotion, labels left EMPTY). Reproducible via SEED.

    python sentiment/src/build_kappa4.py

Outputs (data/annotated/): annotator_A.xlsx … annotator_D.xlsx, overlap_ids.csv (hidden AI key).
"""
from __future__ import annotations
import sys, io
from pathlib import Path
import numpy as np
import pandas as pd
from sklearn.model_selection import StratifiedKFold
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
LABELS = ROOT / "data" / "annotated" / "llm_labels.csv"
OUT = ROOT / "data" / "annotated"
OVERLAP_CSV = OUT / "overlap_ids.csv"

SEED = 42
OVERLAP_N = 450
ANNOTATORS = [("A", "Lamia"), ("B", "Member 2"), ("C", "Member 3"), ("D", "Member 4")]

POLARITY = ["Negative", "Neutral", "Positive"]
EMOTION = ["Anger/Condemnation", "Grief/Sympathy", "Prayer/Religious",
           "Victim-blaming", "Neutral-Other"]
VIEW_COLS = ["comment_id", "platform", "abuse_type", "Comment"]
HEADERS = VIEW_COLS + ["polarity", "emotion", "notes"]


def build_xlsx(df: pd.DataFrame, path: Path, who: str) -> None:
    wb = Workbook()
    ins = wb.active; ins.title = "Instructions"
    for i, (txt, bold) in enumerate([
            (f"Annotation sheet — {who}", True), ("", False),
            ("For EACH row: pick a polarity AND an emotion from the dropdowns.", False),
            ("Label the COMMENT TEXT only. Ignore abuse_type (it is the video topic).", False),
            ("Label INDEPENDENTLY — do not discuss with the other annotators.", False),
            ("The first 450 rows are shared by all 4 (for agreement) — label them like any other.", False),
            ("Do not edit/reorder comment_id. Add a note for hard cases.", False), ("", False),
            ("polarity:  Negative | Neutral | Positive", False),
            ("emotion:   Anger/Condemnation | Grief/Sympathy | Prayer/Religious | "
             "Victim-blaming | Neutral-Other", False), ("", False),
            ("Full rules: sentiment/annotation_guideline.md", False)], start=1):
        c = ins.cell(row=i, column=1, value=txt)
        if bold:
            c.font = Font(bold=True, size=13)
    ins.column_dimensions["A"].width = 100

    ws = wb.create_sheet("Annotate")
    fill = PatternFill("solid", fgColor="DCE6F1")
    for j, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=j, value=h); c.font = Font(bold=True); c.fill = fill
    for _, row in df.iterrows():
        ws.append([row.get(c, "") for c in VIEW_COLS] + ["", "", ""])
    widths = {"comment_id": 10, "platform": 10, "abuse_type": 14, "Comment": 80,
              "polarity": 16, "emotion": 20, "notes": 24}
    for j, h in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = widths.get(h, 14)
    cc = get_column_letter(HEADERS.index("Comment") + 1)
    for r in range(2, ws.max_row + 1):
        ws[f"{cc}{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    n = ws.max_row
    pcol = get_column_letter(HEADERS.index("polarity") + 1)
    ecol = get_column_letter(HEADERS.index("emotion") + 1)
    dp = DataValidation(type="list", formula1='"%s"' % ",".join(POLARITY), allow_blank=True)
    de = DataValidation(type="list", formula1='"%s"' % ",".join(EMOTION), allow_blank=True)
    ws.add_data_validation(dp); ws.add_data_validation(de)
    dp.add(f"{pcol}2:{pcol}{n}"); de.add(f"{ecol}2:{ecol}{n}")
    wb.save(path)


def main() -> None:
    df = pd.read_csv(LABELS, dtype=str, keep_default_na=False)
    OUT.mkdir(parents=True, exist_ok=True)

    # keep the existing 450 overlap if present (continuity), else stratify a fresh one
    if OVERLAP_CSV.exists():
        ov_ids = set(pd.read_csv(OVERLAP_CSV, dtype=str)["comment_id"])
        overlap = df[df["comment_id"].isin(ov_ids)]
    else:
        from sklearn.model_selection import train_test_split
        overlap, _ = train_test_split(df, train_size=OVERLAP_N, random_state=SEED,
                                      stratify=df["emotion"])
        ov_ids = set(overlap["comment_id"])
    rest = df[~df["comment_id"].isin(ov_ids)].reset_index(drop=True)

    # 4-way emotion-stratified split of the rest
    skf = StratifiedKFold(n_splits=4, shuffle=True, random_state=SEED)
    folds = [rest.iloc[idx] for _, idx in skf.split(rest, rest["emotion"])]

    print(f"overlap (all 4): {len(overlap)} | rest split into 4: "
          f"{[len(f) for f in folds]}")
    for (code, who), fold in zip(ANNOTATORS, folds):
        sheet = pd.concat([overlap, fold]).sample(frac=1, random_state=SEED).reset_index(drop=True)
        build_xlsx(sheet, OUT / f"annotator_{code}.xlsx", f"Annotator {code} ({who})")
        print(f"  annotator_{code}.xlsx: {len(sheet)}  (450 overlap + {len(fold)} unique)")

    key = overlap[["comment_id", "polarity", "emotion"]].rename(
        columns={"polarity": "ai_polarity", "emotion": "ai_emotion"})
    key.to_csv(OVERLAP_CSV, index=False, encoding="utf-8")
    print(f"\noverlap_ids.csv (hidden AI key): {len(key)}")
    print("Coverage: every comment labelled ≥1× (450 labelled 4×). Full human-annotated dataset.")


if __name__ == "__main__":
    main()
